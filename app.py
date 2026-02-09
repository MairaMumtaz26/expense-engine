"""
QBO Expense Engine - Streamlit Dashboard
Interactive expense classification, apportionment, and export.
"""
import streamlit as st
import pandas as pd
import os
import sys
import json
from datetime import datetime

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import ACCOUNT_MAP, EXPENSE_CATEGORIES, PL_BRANDS, ALL_BRANDS
from parser import parse_all_files
from classifier import classify_dataframe, load_learned_rules, save_learned_rules
from apportionment import (
    load_shared_rules, save_shared_rules,
    load_apportionment, save_apportionment,
    get_default_apportionment, apportion_transaction,
    load_txn_apportionment, save_txn_apportionment, apportion_by_splits,
)
from fx import load_fx_rates, save_fx_rates, convert_cad_to_usd, get_required_months

st.set_page_config(page_title="QBO Expense Engine", page_icon="💰", layout="wide")

# ============================================================
# PERSISTENT STATE — survives browser refresh
# ============================================================
STATE_DIR = "data"
RAW_DF_FILE = os.path.join(STATE_DIR, "_raw_df.pkl")
CLASSIFIED_DF_FILE = os.path.join(STATE_DIR, "_classified_df.pkl")
STEP_FILE = os.path.join(STATE_DIR, "_step.json")
DISMISSED_FILE = os.path.join(STATE_DIR, "_dismissed_inconsistencies.json")


def save_state_to_disk():
    """Persist current progress to disk."""
    os.makedirs(STATE_DIR, exist_ok=True)
    if st.session_state.raw_df is not None:
        st.session_state.raw_df.to_pickle(RAW_DF_FILE)
    if st.session_state.classified_df is not None:
        st.session_state.classified_df.to_pickle(CLASSIFIED_DF_FILE)
    with open(STEP_FILE, "w") as f:
        json.dump({"step": st.session_state.step}, f)
    if "dismissed_inconsistencies" in st.session_state and st.session_state.dismissed_inconsistencies:
        with open(DISMISSED_FILE, "w") as f:
            json.dump(list(st.session_state.dismissed_inconsistencies), f)


def load_state_from_disk():
    """Restore progress from disk after browser refresh."""
    if os.path.exists(RAW_DF_FILE):
        st.session_state.raw_df = pd.read_pickle(RAW_DF_FILE)
    if os.path.exists(CLASSIFIED_DF_FILE):
        st.session_state.classified_df = pd.read_pickle(CLASSIFIED_DF_FILE)
    if os.path.exists(STEP_FILE):
        with open(STEP_FILE, "r") as f:
            st.session_state.step = json.load(f).get("step", 1)
    if os.path.exists(DISMISSED_FILE):
        with open(DISMISSED_FILE, "r") as f:
            st.session_state.dismissed_inconsistencies = set(json.load(f))


def clear_state_from_disk():
    """Remove persisted state files."""
    for f in [RAW_DF_FILE, CLASSIFIED_DF_FILE, STEP_FILE, DISMISSED_FILE]:
        if os.path.exists(f):
            os.remove(f)


# ============================================================
# SESSION STATE INITIALIZATION
# ============================================================
if "raw_df" not in st.session_state:
    st.session_state.raw_df = None
if "classified_df" not in st.session_state:
    st.session_state.classified_df = None
if "step" not in st.session_state:
    st.session_state.step = 1
if "unidentified_files" not in st.session_state:
    st.session_state.unidentified_files = []

# Auto-restore from disk on fresh load (browser refresh)
if st.session_state.raw_df is None and os.path.exists(RAW_DF_FILE):
    load_state_from_disk()


def reset_pipeline():
    st.session_state.raw_df = None
    st.session_state.classified_df = None
    st.session_state.step = 1
    st.session_state.unidentified_files = []
    clear_state_from_disk()


def search_filter(df, query):
    """Case-insensitive partial keyword search. Multiple words = AND match."""
    if not query or not query.strip() or df.empty:
        return df
    terms = query.strip().lower().split()
    search_cols = [c for c in ["name", "memo", "head", "sub_head", "description", "brand", "qbo_account"] if c in df.columns]
    mask = pd.Series(True, index=df.index)
    for term in terms:
        term_mask = pd.Series(False, index=df.index)
        for col in search_cols:
            term_mask |= df[col].astype(str).str.lower().str.contains(term, na=False, regex=False)
        mask &= term_mask
    return df[mask]


def add_transaction_form():
    """Render an 'Add Transaction' form. Works at any step."""
    df = st.session_state.classified_df
    if df is None:
        return
    step = st.session_state.step
    with st.expander("➕ Add New Transaction", expanded=False):
        with st.form(f"add_txn_step{step}", clear_on_submit=True):
            fc1, fc2 = st.columns(2)
            with fc1:
                new_date = st.date_input("Date", value=datetime.now())
                new_name = st.text_input("Name (Vendor)")
                new_memo = st.text_input("Memo / Description")
            with fc2:
                new_amount = st.number_input("Amount (LCY)", value=0.0, step=0.01, format="%.2f")
                new_currency = st.selectbox("Currency", ["USD", "CAD"])
                existing_brands = sorted(df["brand"].dropna().unique().tolist())
                if existing_brands:
                    new_brand = st.selectbox("Brand", existing_brands)
                else:
                    new_brand = st.text_input("Brand")

            fc3, fc4 = st.columns(2)
            with fc3:
                existing_accounts = sorted(df["qbo_account"].dropna().unique().tolist())
                if existing_accounts:
                    new_account = st.selectbox("QBO Account", existing_accounts)
                else:
                    new_account = st.text_input("QBO Account")
                head_opts = sorted(EXPENSE_CATEGORIES.keys())
                new_head = st.selectbox("Head", ["-- Select --"] + head_opts)
            with fc4:
                new_sub = st.text_input("Sub-head")
                if "shared_tag" in df.columns:
                    new_tag = st.selectbox("Shared / Direct", ["direct", "shared"])
                else:
                    new_tag = "direct"

            submitted = st.form_submit_button("➕ Add Transaction", type="primary", use_container_width=True)

            if submitted:
                if not new_name.strip() or new_amount == 0 or new_head == "-- Select --":
                    st.warning("Please fill in Name, a non-zero Amount, and Head.")
                else:
                    new_idx = int(df.index.max()) + 1 if len(df) > 0 else 0
                    description = f"{new_name} {new_memo}".strip()

                    # Build row with defaults for every column
                    new_row = {}
                    for col in df.columns:
                        if pd.api.types.is_datetime64_any_dtype(df[col]):
                            new_row[col] = pd.NaT
                        elif df[col].dtype == object:
                            new_row[col] = ""
                        else:
                            new_row[col] = 0

                    new_row.update({
                        "date": pd.Timestamp(new_date),
                        "name": new_name.strip(),
                        "memo": new_memo.strip(),
                        "amount_lcy": new_amount,
                        "currency": new_currency,
                        "brand": new_brand,
                        "qbo_account": new_account if new_account else "",
                        "qbo_group": "",
                        "account_type": "expense",
                        "description": description,
                        "head": new_head,
                        "sub_head": new_sub.strip() if new_sub.strip() else new_head,
                        "match_confidence": "manual",
                    })
                    if "shared_tag" in df.columns:
                        new_row["shared_tag"] = new_tag

                    new_df_row = pd.DataFrame([new_row], index=[new_idx])
                    st.session_state.classified_df = pd.concat([df, new_df_row])

                    # Also add to raw_df
                    if st.session_state.raw_df is not None:
                        raw = st.session_state.raw_df
                        raw_row = {}
                        for col in raw.columns:
                            if col in new_row:
                                raw_row[col] = new_row[col]
                            elif pd.api.types.is_datetime64_any_dtype(raw[col]):
                                raw_row[col] = pd.NaT
                            elif raw[col].dtype == object:
                                raw_row[col] = ""
                            else:
                                raw_row[col] = 0
                        st.session_state.raw_df = pd.concat([raw, pd.DataFrame([raw_row], index=[new_idx])])

                    save_state_to_disk()
                    st.success(f"✅ Added: {new_name} | {new_head} > {new_row['sub_head']} | ${new_amount:,.2f}")
                    st.rerun()


def delete_transactions(indices):
    """Remove transactions by index from both dataframes. Preserves all other saved data."""
    if st.session_state.classified_df is not None:
        st.session_state.classified_df = st.session_state.classified_df.drop(indices, errors="ignore")
    if st.session_state.raw_df is not None:
        st.session_state.raw_df = st.session_state.raw_df.drop(indices, errors="ignore")
    # Clean up transaction-level apportionment for deleted rows
    txn_app = load_txn_apportionment()
    changed = False
    for idx in indices:
        if str(idx) in txn_app:
            del txn_app[str(idx)]
            changed = True
    if changed:
        save_txn_apportionment(txn_app)
    save_state_to_disk()


# ============================================================
# SIDEBAR - Navigation
# ============================================================
st.sidebar.title("💰 QBO Expense Engine")
st.sidebar.markdown("---")

steps = {
    1: "📁 Upload GL Files",
    2: "🏷️ Classify Expenses",
    3: "🔄 Shared / Direct",
    4: "📊 Apportionment %",
    5: "💱 FX Rates",
    6: "📤 Export CSV",
}

for step_num, step_name in steps.items():
    if step_num <= st.session_state.step:
        if st.sidebar.button(step_name, key=f"nav_{step_num}", use_container_width=True):
            st.session_state.step = step_num
    else:
        st.sidebar.button(step_name, key=f"nav_{step_num}", disabled=True, use_container_width=True)

st.sidebar.markdown("---")
if st.sidebar.button("🔄 Start Over", use_container_width=True):
    reset_pipeline()
    st.rerun()

# Show stats if data loaded
if st.session_state.raw_df is not None:
    df = st.session_state.raw_df
    st.sidebar.markdown("### 📈 Summary")
    st.sidebar.metric("Total Transactions", len(df))
    st.sidebar.metric("QBO Accounts", df["qbo_account"].nunique())
    st.sidebar.metric("Total Amount (LCY)", f"${df['amount_lcy'].sum():,.2f}")


# ============================================================
# STEP 1: Upload GL Files
# ============================================================
if st.session_state.step == 1:
    st.title("📁 Step 1: Upload GL Files")
    st.markdown("Upload your QBO General Ledger Excel files. The system auto-identifies each account from the filename.")

    uploaded_files = st.file_uploader(
        "Drop all GL files here",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
    )

    if uploaded_files:
        # Save temp files and parse
        temp_paths = []
        os.makedirs("data/temp", exist_ok=True)
        for uf in uploaded_files:
            temp_path = f"data/temp/{uf.name}"
            with open(temp_path, "wb") as f:
                f.write(uf.getbuffer())
            temp_paths.append(temp_path)

        with st.spinner("Parsing GL files..."):
            df, unidentified = parse_all_files(temp_paths)

        if unidentified:
            st.warning(f"⚠️ Could not identify: {', '.join(unidentified)}")
            st.session_state.unidentified_files = unidentified

        if not df.empty:
            st.session_state.raw_df = df
            st.success(f"✅ Loaded {len(df)} transactions from {df['qbo_account'].nunique()} accounts")

            # Show breakdown
            summary = df.groupby(["qbo_account", "brand", "currency"]).agg(
                transactions=("amount_lcy", "count"),
                total_amount=("amount_lcy", "sum"),
            ).reset_index()
            st.dataframe(summary, use_container_width=True)

            if st.button("▶️ Proceed to Classification", type="primary", use_container_width=True):
                st.session_state.step = 2
                st.rerun()
        else:
            st.error("No transactions found in uploaded files.")


# ============================================================
# STEP 2: Classify Expenses
# ============================================================
elif st.session_state.step == 2:
    st.title("🏷️ Step 2: Classify Expenses")

    # Only classify once — use cached result after that
    if st.session_state.classified_df is None:
        df = st.session_state.raw_df.copy()
        with st.spinner("Classifying transactions..."):
            df = classify_dataframe(df)
            st.session_state.classified_df = df

    df = st.session_state.classified_df

    add_transaction_form()

    search_q2 = st.text_input("🔍 Search transactions", placeholder="Type to filter by name, memo, head, brand...", key="search_s2")

    # Button to re-classify from scratch (picks up new learned rules)
    if st.button("🔄 Re-classify All", help="Re-run classification from scratch using saved rules"):
        old_df = st.session_state.classified_df
        fresh = st.session_state.raw_df.copy()
        fresh = classify_dataframe(fresh)
        # Preserve shared_tag from previous Step 3 work
        if old_df is not None and "shared_tag" in old_df.columns:
            fresh["shared_tag"] = old_df["shared_tag"]
        st.session_state.classified_df = fresh
        st.rerun()

    # Split into 3 buckets
    clean = df[df["match_confidence"].isin(["keyword", "qbo_group", "learned", "manual"])]
    discrepancy = df[df["match_confidence"] == "discrepancy"]
    unmatched = df[df["match_confidence"] == "unmatched"]

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("✅ Clean", len(clean))
    col2.metric("🟠 Discrepancy", len(discrepancy))
    col3.metric("❌ Unmatched", len(unmatched))
    col4.metric("Match Rate", f"{(len(clean)+len(discrepancy))/max(len(df),1)*100:.1f}%")

    # Apply search filter for display (metrics above show totals)
    if search_q2:
        clean = search_filter(clean, search_q2)
        unmatched = search_filter(unmatched, search_q2)
        # discrepancy filtered below after auto-save pass runs on full set

    head_options = sorted(EXPENSE_CATEGORIES.keys())

    # ---- SECTION 1: Clean matches (editable, sorted by Head → Sub-head) ----
    st.subheader("✅ Auto-Classified Transactions (sorted by Head / Sub-head — click to edit)")

    # Filter by Head
    if not clean.empty:
        all_heads_in_data = sorted(clean["head"].dropna().unique().tolist())
        filter_head = st.selectbox("Filter by Head", ["All"] + all_heads_in_data, key="filter_head_clean")

        filtered_clean = clean if filter_head == "All" else clean[clean["head"] == filter_head]

        # Show existing sub-heads as reference for consistency
        _all_classified = df[df["sub_head"].notna() & (df["sub_head"] != "")]
        if filter_head != "All":
            _ref_heads = [filter_head]
        else:
            _ref_heads = all_heads_in_data
        ref_data = {}
        for h in _ref_heads:
            subs = sorted(_all_classified[_all_classified["head"] == h]["sub_head"].unique().tolist())
            subs = [s for s in subs if s.lower() != h.lower()]
            if subs:
                ref_data[h] = subs
        if ref_data:
            with st.expander("📋 Existing Sub-heads (reference for consistent naming)", expanded=False):
                for h, subs in ref_data.items():
                    st.markdown(f"**{h}:** {', '.join(subs)}")

        display_cols = ["head", "sub_head", "name", "memo", "brand", "amount_lcy", "match_confidence"]
        edit_df = filtered_clean[display_cols].sort_values(["head", "sub_head", "name"]).copy()
        edit_df = edit_df.reset_index()  # keep original index for write-back
        edit_df["🗑️"] = False  # delete column

        edited = st.data_editor(
            edit_df,
            column_config={
                "index": None,
                "🗑️": st.column_config.CheckboxColumn("🗑️", default=False),
                "head": st.column_config.SelectboxColumn("Head", options=head_options, required=True),
                "sub_head": st.column_config.TextColumn("Sub-head"),
                "name": st.column_config.TextColumn("Name"),
                "memo": st.column_config.TextColumn("Memo", disabled=True),
                "brand": st.column_config.TextColumn("Brand", disabled=True),
                "amount_lcy": st.column_config.NumberColumn("Amount", format="$%.2f"),
                "match_confidence": st.column_config.TextColumn("Confidence", disabled=True),
            },
            column_order=["head", "sub_head", "name", "memo", "brand", "amount_lcy", "match_confidence", "🗑️"],
            use_container_width=True,
            height=500,
            num_rows="fixed",
            key="classify_editor",
        )

        if st.button("💾 Save Reclassifications", key="save_reclassify"):
            learned_rules = load_learned_rules()
            changes = 0
            skipped = 0
            deleted = 0
            for _, erow in edited.iterrows():
                orig_idx = erow["index"]
                # Handle deletions
                if erow.get("🗑️", False):
                    df = df.drop(orig_idx)
                    if st.session_state.raw_df is not None and orig_idx in st.session_state.raw_df.index:
                        st.session_state.raw_df = st.session_state.raw_df.drop(orig_idx)
                    deleted += 1
                    continue
                orig_head = df.at[orig_idx, "head"]
                orig_sub = df.at[orig_idx, "sub_head"]
                new_head = erow["head"]
                new_sub = erow["sub_head"]
                # Skip rows where head is blank — leave as-is
                if not new_head or (isinstance(new_head, str) and not new_head.strip()):
                    skipped += 1
                    continue
                if not new_sub or (isinstance(new_sub, str) and not new_sub.strip()):
                    new_sub = orig_sub  # keep original sub if blank
                if new_head != orig_head or new_sub != orig_sub:
                    df.at[orig_idx, "head"] = new_head
                    df.at[orig_idx, "sub_head"] = new_sub
                    df.at[orig_idx, "match_confidence"] = "manual"
                    desc_key = df.at[orig_idx, "description"].strip().lower()
                    learned_rules[desc_key] = {"head": new_head, "sub_head": new_sub}
                    changes += 1
                # Check for amount changes
                new_amount = float(erow["amount_lcy"])
                if abs(new_amount - df.at[orig_idx, "amount_lcy"]) > 0.001:
                    df.at[orig_idx, "amount_lcy"] = new_amount
                    if st.session_state.raw_df is not None and orig_idx in st.session_state.raw_df.index:
                        st.session_state.raw_df.at[orig_idx, "amount_lcy"] = new_amount
                    changes += 1
            if changes > 0 or deleted > 0:
                save_learned_rules(learned_rules)
                st.session_state.classified_df = df
                parts = []
                if changes:
                    parts.append(f"{changes} reclassified")
                if deleted:
                    parts.append(f"{deleted} removed")
                if skipped:
                    parts.append(f"{skipped} skipped (blank head)")
                msg = f"✅ {', '.join(parts)}. Rules remembered for next cycle."
                st.success(msg)
                # Scroll back to save button instead of jumping to top
                import streamlit.components.v1 as components
                components.html("""
                    <script>
                    setTimeout(function() {
                        const btns = window.parent.document.querySelectorAll('button');
                        for (const btn of btns) {
                            if (btn.textContent.includes('Save Reclassifications')) {
                                btn.scrollIntoView({behavior: 'smooth', block: 'center'});
                                break;
                            }
                        }
                    }, 300);
                    </script>
                """, height=0)
            else:
                st.info("No changes detected.")
    else:
        st.info("No clean matches yet.")

    # ---- SECTION 2: Discrepancies (sub_head == head) — ORANGE ----
    if not discrepancy.empty:
        st.markdown("---")
        st.markdown(
            '<div style="background-color:#FFF3E0; border-left:5px solid #FF9800; padding:12px; '
            'border-radius:4px; margin-bottom:16px;">'
            '<h3 style="color:#E65100; margin:0;">🟠 Discrepancies — Sub-head = Head '
            f'({len(discrepancy)} transactions)</h3>'
            '<p style="color:#BF360C; margin:4px 0 0 0;">Head was identified but no specific '
            'Sub-head keyword found in Name or Memo. Please select the correct Sub-head below.</p>'
            '</div>',
            unsafe_allow_html=True,
        )

        learned_rules = load_learned_rules()

        # Undo history: {idx: {"desc_key": ..., "old_sub": ..., "old_confidence": ...}}
        if "disc_undo" not in st.session_state:
            st.session_state.disc_undo = {}

        # Track which field was just saved for auto-navigate
        if "disc_last_saved_idx" not in st.session_state:
            st.session_state.disc_last_saved_idx = None

        # Build suggestion index: for each Head, collect sub-heads used in learned rules
        # Also index by description words for similarity matching
        _sub_by_head = {}   # head -> set of sub-heads
        _sub_by_word = {}   # word -> set of sub-heads
        for desc_k, rule in learned_rules.items():
            h, s = rule["head"], rule["sub_head"]
            _sub_by_head.setdefault(h, set()).add(s)
            for w in desc_k.split():
                if len(w) > 3:  # skip short words
                    _sub_by_word.setdefault(w, set()).add(s)

        def get_suggestions(head, description, name, memo):
            """Return ranked list of suggested sub-heads."""
            suggestions = []
            # 1. Sub-heads from similar descriptions (shared words)
            words = set(f"{name} {memo} {description}".lower().split())
            word_scores = {}
            for w in words:
                if w in _sub_by_word:
                    for s in _sub_by_word[w]:
                        word_scores[s] = word_scores.get(s, 0) + 1
            # Sort by how many words matched (most similar first)
            ranked = sorted(word_scores.items(), key=lambda x: -x[1])
            suggestions.extend([s for s, _ in ranked])
            # 2. All sub-heads previously used for same Head
            if head in _sub_by_head:
                for s in sorted(_sub_by_head[head]):
                    if s not in suggestions:
                        suggestions.append(s)
            # Remove head name itself
            suggestions = [s for s in suggestions if s.lower() != head.lower()]
            return suggestions

        # Auto-save pass: check text inputs AND selectbox inputs
        for idx, row in discrepancy.iterrows():
            # Check text input
            sub_key = f"disc_sub_{idx}"
            if sub_key in st.session_state:
                typed_val = st.session_state[sub_key]
                if typed_val and typed_val.strip() and df.at[idx, "match_confidence"] == "discrepancy":
                    st.session_state.disc_last_saved_idx = idx
                    desc_key = row["description"].strip().lower()
                    st.session_state.disc_undo[idx] = {
                        "desc_key": desc_key,
                        "old_sub": row["head"],
                        "old_confidence": "discrepancy",
                    }
                    learned_rules[desc_key] = {"head": row["head"], "sub_head": typed_val.strip()}
                    save_learned_rules(learned_rules)
                    df.at[idx, "sub_head"] = typed_val.strip()
                    df.at[idx, "match_confidence"] = "manual"
                    st.session_state.classified_df = df
                    _sub_by_head.setdefault(row["head"], set()).add(typed_val.strip())
                    continue
            # Check suggestion selectbox
            sug_key = f"disc_sug_{idx}"
            if sug_key in st.session_state:
                sug_val = st.session_state[sug_key]
                if sug_val and sug_val != "-- Suggestions --" and sug_val != "Type custom below..." and df.at[idx, "match_confidence"] == "discrepancy":
                    st.session_state.disc_last_saved_idx = idx
                    desc_key = row["description"].strip().lower()
                    st.session_state.disc_undo[idx] = {
                        "desc_key": desc_key,
                        "old_sub": row["head"],
                        "old_confidence": "discrepancy",
                    }
                    learned_rules[desc_key] = {"head": row["head"], "sub_head": sug_val}
                    save_learned_rules(learned_rules)
                    df.at[idx, "sub_head"] = sug_val
                    df.at[idx, "match_confidence"] = "manual"
                    st.session_state.classified_df = df
                    _sub_by_head.setdefault(row["head"], set()).add(sug_val)

        # Collect all discrepancy + recently saved indices for display in order
        disc_and_saved = df[df["match_confidence"].isin(["discrepancy", "manual"]) & (
            (df["match_confidence"] == "discrepancy") |
            (df.index.isin(st.session_state.disc_undo.keys()))
        )]
        if search_q2:
            disc_and_saved = search_filter(disc_and_saved, search_q2)

        # Build ordered list of indices for auto-navigate
        all_disc_indices = list(disc_and_saved.index)
        last_saved = st.session_state.disc_last_saved_idx
        # Find position of next field after last saved
        focus_position = 0
        if last_saved is not None and last_saved in all_disc_indices:
            saved_pos = all_disc_indices.index(last_saved)
            focus_position = saved_pos + 1  # next after saved

        field_counter = 0
        has_remaining = False

        if not disc_and_saved.empty:
            for acct_name, acct_group in disc_and_saved.groupby("qbo_account"):
                pending = len(acct_group[acct_group["match_confidence"] == "discrepancy"])
                done = len(acct_group[acct_group["match_confidence"] == "manual"])
                label = f"📂 {acct_name} — {pending} pending"
                if done:
                    label += f", {done} saved"
                with st.expander(label, expanded=True):
                    for idx, row in acct_group.iterrows():
                        memo_text = row["memo"] if row["memo"] else "—"
                        currency = row["currency"] if "currency" in row.index else "USD"
                        is_saved = (row["match_confidence"] == "manual" and idx in st.session_state.disc_undo)

                        if is_saved:
                            # Saved row — show sub-head with undo
                            c1, ca, c2, c3 = st.columns([2.5, 0.8, 0.7, 0.3])
                            with c1:
                                st.caption(f"✅ Head: {row['head']}  |  Sub-head: **{row['sub_head']}**  |  Memo: {memo_text}")
                            with ca:
                                _amt = st.number_input("Amt", value=float(df.at[idx, "amount_lcy"]), step=0.01, format="%.2f", key=f"amt_ds_{idx}", label_visibility="collapsed")
                                if abs(_amt - df.at[idx, "amount_lcy"]) > 0.001:
                                    df.at[idx, "amount_lcy"] = _amt
                                    if st.session_state.raw_df is not None and idx in st.session_state.raw_df.index:
                                        st.session_state.raw_df.at[idx, "amount_lcy"] = _amt
                                    st.session_state.classified_df = df
                            with c2:
                                if st.button("↩️ Undo", key=f"disc_undo_{idx}"):
                                    undo = st.session_state.disc_undo.pop(idx)
                                    df.at[idx, "sub_head"] = undo["old_sub"]
                                    df.at[idx, "match_confidence"] = undo["old_confidence"]
                                    if undo["desc_key"] in learned_rules:
                                        del learned_rules[undo["desc_key"]]
                                        save_learned_rules(learned_rules)
                                    st.session_state.classified_df = df
                                    st.rerun()
                            with c3:
                                if st.button("🗑️", key=f"disc_del_{idx}", help="Delete transaction"):
                                    delete_transactions([idx])
                                    st.rerun()
                        else:
                            # Pending row — suggestions + text input
                            has_remaining = True
                            suggestions = get_suggestions(row["head"], row["description"], row["name"], memo_text)
                            c1, ca, c2, c3 = st.columns([1.5, 0.8, 2, 0.3])
                            with c1:
                                st.caption(f"Head: {row['head']}  |  Memo: {memo_text}")
                            with ca:
                                _amt = st.number_input("Amt", value=float(df.at[idx, "amount_lcy"]), step=0.01, format="%.2f", key=f"amt_dp_{idx}", label_visibility="collapsed")
                                if abs(_amt - df.at[idx, "amount_lcy"]) > 0.001:
                                    df.at[idx, "amount_lcy"] = _amt
                                    if st.session_state.raw_df is not None and idx in st.session_state.raw_df.index:
                                        st.session_state.raw_df.at[idx, "amount_lcy"] = _amt
                                    st.session_state.classified_df = df
                            with c2:
                                if suggestions:
                                    st.selectbox(
                                        "Suggested sub-heads (select to save)",
                                        ["-- Suggestions --"] + suggestions + ["Type custom below..."],
                                        key=f"disc_sug_{idx}",
                                    )
                                st.text_input(
                                    "Or type custom (press Enter to save)",
                                    value="",
                                    placeholder="Type sub-head & press Enter...",
                                    key=f"disc_sub_{idx}",
                                )
                            with c3:
                                if st.button("🗑️", key=f"disc_del_p_{idx}", help="Delete transaction"):
                                    delete_transactions([idx])
                                    st.rerun()
                        field_counter += 1

            if has_remaining and last_saved is not None:
                # Only auto-focus next field if a discrepancy was just saved
                import streamlit.components.v1 as components
                components.html(f"""
                    <script>
                    function focusNext() {{
                        const inputs = window.parent.document.querySelectorAll('input[placeholder*="sub-head"]');
                        let emptyInputs = [];
                        for (const inp of inputs) {{
                            if (!inp.value) emptyInputs.push(inp);
                        }}
                        if (emptyInputs.length > 0) {{
                            emptyInputs[0].focus();
                            emptyInputs[0].scrollIntoView({{behavior: 'smooth', block: 'center'}});
                        }}
                    }}
                    setTimeout(focusNext, 300);
                    </script>
                """, height=0)
                # Reset so it doesn't trigger again on next rerun
                st.session_state.disc_last_saved_idx = None

        if not has_remaining and disc_and_saved.empty:
            st.success("All discrepancies resolved!")

    # ---- SECTION 2b: Name inconsistencies (same Name, different Sub-heads) ----
    if "dismissed_inconsistencies" not in st.session_state:
        st.session_state.dismissed_inconsistencies = set()

    classified_with_sub = df[df["sub_head"].notna() & (df["sub_head"] != "") & (df["name"].notna()) & (df["name"] != "")]
    if not classified_with_sub.empty:
        name_subs = classified_with_sub.groupby("name")["sub_head"].apply(lambda x: list(x.unique())).reset_index()
        inconsistent = name_subs[name_subs["sub_head"].apply(len) > 1]
        # Filter out dismissed ones
        inconsistent = inconsistent[~inconsistent["name"].isin(st.session_state.dismissed_inconsistencies)]
        if not inconsistent.empty:
            st.markdown("---")
            st.markdown(
                '<div style="background-color:#FCE4EC; border-left:5px solid #E91E63; padding:12px; '
                'border-radius:4px; margin-bottom:16px;">'
                f'<h3 style="color:#880E4F; margin:0;">⚠️ Name Inconsistencies ({len(inconsistent)} names with mixed Sub-heads)</h3>'
                '<p style="color:#AD1457; margin:4px 0 0 0;">Same Name is classified under different Sub-heads. Mark OK if intentional.</p>'
                '</div>',
                unsafe_allow_html=True,
            )
            learned_rules_inc = load_learned_rules()
            for _, inc_row in inconsistent.iterrows():
                vendor_name = inc_row["name"]
                sub_list = inc_row["sub_head"]
                vendor_txns = classified_with_sub[classified_with_sub["name"] == vendor_name]
                col_title, col_ok = st.columns([4, 1])
                with col_title:
                    expander_open = st.expander(f"⚠️ {vendor_name} — used as: {', '.join(sub_list)} ({len(vendor_txns)} txns)", expanded=False)
                with col_ok:
                    if st.button("✅ OK", key=f"inc_ok_{vendor_name}"):
                        st.session_state.dismissed_inconsistencies.add(vendor_name)
                        st.rerun()
                with expander_open:
                    for idx, row in vendor_txns.iterrows():
                        memo_text = row["memo"] if row["memo"] else "—"
                        currency = row["currency"] if "currency" in row.index else "USD"
                        c1, c2, c3 = st.columns([3, 2, 1])
                        with c1:
                            st.caption(f"Head: {row['head']}  |  Sub-head: {row['sub_head']}  |  Memo: {memo_text}  |  {currency} {row['amount_lcy']:,.2f}")
                        with c2:
                            new_sub = st.text_input(
                                "Fix Sub-head",
                                value=row["sub_head"],
                                key=f"inc_sub_{idx}",
                            )
                        with c3:
                            if new_sub != row["sub_head"]:
                                if st.button("✅ Fix", key=f"inc_save_{idx}"):
                                    desc_key = row["description"].strip().lower()
                                    learned_rules_inc[desc_key] = {"head": row["head"], "sub_head": new_sub}
                                    save_learned_rules(learned_rules_inc)
                                    df.at[idx, "sub_head"] = new_sub
                                    df.at[idx, "match_confidence"] = "manual"
                                    st.session_state.classified_df = df
                                    st.rerun()

    # ---- SECTION 3: Unmatched (no head at all) — RED ----
    if not unmatched.empty:
        st.markdown("---")
        st.subheader(f"❌ Unmatched Transactions ({len(unmatched)}) — Classify Now")
        st.markdown("No Head or Sub-head could be identified. Select both below.")

        learned_rules = load_learned_rules()
        head_options_with_blank = ["-- Select --"] + head_options

        for idx, row in unmatched.iterrows():
            with st.expander(
                f"🔴 {row['brand']} | {row['name']} | {row['memo'][:60]} | ${row['amount_lcy']:,.2f}",
                expanded=True,
            ):
                c1, c2 = st.columns(2)
                with c1:
                    st.text(f"QBO Group: {row['qbo_group']}")
                    st.text(f"Description: {row['description'][:100]}")
                    st.text(f"Date: {row['date'].strftime('%Y-%m-%d') if hasattr(row['date'], 'strftime') else row['date']}")
                    _amt = st.number_input("Amount", value=float(df.at[idx, "amount_lcy"]), step=0.01, format="%.2f", key=f"amt_um_{idx}")
                    if abs(_amt - df.at[idx, "amount_lcy"]) > 0.001:
                        df.at[idx, "amount_lcy"] = _amt
                        if st.session_state.raw_df is not None and idx in st.session_state.raw_df.index:
                            st.session_state.raw_df.at[idx, "amount_lcy"] = _amt
                        st.session_state.classified_df = df
                    if st.button("🗑️ Delete", key=f"del_unmatched_{idx}"):
                        delete_transactions([idx])
                        st.rerun()

                with c2:
                    selected_head = st.selectbox(
                        "Head Category",
                        head_options_with_blank,
                        key=f"head_{idx}",
                    )
                    if selected_head != "-- Select --":
                        sub_opts = [kw for kw in EXPENSE_CATEGORIES.get(selected_head, []) if kw.lower() != selected_head.lower()]
                        if not sub_opts:
                            sub_opts = EXPENSE_CATEGORIES.get(selected_head, [selected_head])
                        selected_sub = st.selectbox(
                            "Sub-head",
                            sub_opts,
                            key=f"sub_{idx}",
                        )
                    else:
                        selected_sub = None

                    if selected_head != "-- Select --" and selected_sub:
                        if st.button(f"✅ Save Rule", key=f"save_{idx}"):
                            desc_key = row["description"].strip().lower()
                            learned_rules[desc_key] = {
                                "head": selected_head,
                                "sub_head": selected_sub,
                            }
                            save_learned_rules(learned_rules)
                            df.at[idx, "head"] = selected_head
                            df.at[idx, "sub_head"] = selected_sub
                            df.at[idx, "match_confidence"] = "learned"
                            st.session_state.classified_df = df
                            st.success(f"Saved! '{row['description'][:40]}...' → {selected_head} > {selected_sub}")
                            st.rerun()

    st.markdown("---")
    needs_action = df[df["match_confidence"].isin(["unmatched", "discrepancy"])]
    if needs_action.empty:
        st.success("🎉 All transactions classified with specific Sub-heads!")
    else:
        disc_left = len(df[df["match_confidence"] == "discrepancy"])
        unm_left = len(df[df["match_confidence"] == "unmatched"])
        parts = []
        if disc_left:
            parts.append(f"{disc_left} discrepancies")
        if unm_left:
            parts.append(f"{unm_left} unmatched")
        st.info(f"ℹ️ {' + '.join(parts)} still need attention. You can proceed and handle them later.")

    if st.button("▶️ Proceed to Shared/Direct Tagging", type="primary", use_container_width=True):
        st.session_state.classified_df = df
        st.session_state.step = 3
        st.rerun()


# ============================================================
# STEP 3: Shared / Direct Tagging
# ============================================================
elif st.session_state.step == 3:
    st.title("🔄 Step 3: Tag Shared vs Direct Expenses")
    st.markdown("""
    Tag each transaction as **Shared** (apportioned across brands) or **Direct** (stays with this brand).
    Click a button to tag — selection shows green. All transactions stay visible.
    """)
    # Light green for selected buttons
    st.markdown("""
    <style>
    .stButton > button[kind="primary"] {
        background-color: #81C784 !important;
        border-color: #66BB6A !important;
        color: white !important;
    }
    .stButton > button[data-testid="baseButton-primary"] {
        background-color: #81C784 !important;
        border-color: #66BB6A !important;
        color: white !important;
    }
    </style>
    """, unsafe_allow_html=True)

    add_transaction_form()

    search_q3 = st.text_input("🔍 Search transactions", placeholder="Type to filter by name, memo, head, brand...", key="search_s3")

    df = st.session_state.classified_df.copy()
    shared_rules = load_shared_rules()

    if "shared_tag" not in df.columns:
        df["shared_tag"] = ""

    # Apply saved rules to restore previous tags
    if st.button("🔄 Apply Saved Rules", help="Re-apply all previously learned shared/direct tags"):
        applied = 0
        for idx, row in df.iterrows():
            if not df.at[idx, "shared_tag"] in ["shared", "direct"]:
                desc_key = row["description"].strip().lower()
                if desc_key in shared_rules:
                    df.at[idx, "shared_tag"] = shared_rules[desc_key]
                    applied += 1
        st.session_state.classified_df = df
        st.success(f"✅ Restored {applied} tags from saved rules!")
        st.rerun()

    tagged_count = len(df[df["shared_tag"].isin(["shared", "direct"])])
    untagged_count = len(df[~df["shared_tag"].isin(["shared", "direct"])])
    shared_count = len(df[df["shared_tag"] == "shared"])
    direct_count = len(df[df["shared_tag"] == "direct"])

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("✅ Tagged", tagged_count)
    col2.metric("❓ Untagged", untagged_count)
    col3.metric("🔄 Shared", shared_count)
    col4.metric("📌 Direct", direct_count)

    # Bulk tagging by Head category
    st.subheader("⚡ Quick Tag by Category")
    st.markdown("Tag all transactions in a category at once:")

    if not df.empty and "head" in df.columns:
        heads_in_data = sorted(df[df["head"].notna()]["head"].unique())
        for head in heads_in_data:
            head_df = df[df["head"] == head]
            head_shared = len(head_df[head_df["shared_tag"] == "shared"])
            head_direct = len(head_df[head_df["shared_tag"] == "direct"])
            head_untagged = len(head_df[~head_df["shared_tag"].isin(["shared", "direct"])])

            # Status indicator
            if head_untagged == 0:
                status = f"✅ {head_shared} shared, {head_direct} direct"
            else:
                status = f"{head_untagged} untagged, {head_shared} shared, {head_direct} direct"

            with st.expander(f"📂 {head} ({len(head_df)} total) — {status}"):
                st.dataframe(
                    head_df[["brand", "name", "memo", "sub_head", "amount_lcy", "shared_tag"]].sort_values("name"),
                    use_container_width=True,
                    height=200,
                )
                c1, c2 = st.columns(2)
                with c1:
                    if st.button(f"Mark ALL '{head}' as SHARED", key=f"bulk_shared_{head}"):
                        for idx2, row2 in head_df.iterrows():
                            desc_key = row2["description"].strip().lower()
                            shared_rules[desc_key] = "shared"
                            df.at[idx2, "shared_tag"] = "shared"
                        save_shared_rules(shared_rules)
                        st.session_state.classified_df = df
                        st.rerun()
                with c2:
                    if st.button(f"Mark ALL '{head}' as DIRECT", key=f"bulk_direct_{head}"):
                        for idx2, row2 in head_df.iterrows():
                            desc_key = row2["description"].strip().lower()
                            shared_rules[desc_key] = "direct"
                            df.at[idx2, "shared_tag"] = "direct"
                        save_shared_rules(shared_rules)
                        st.session_state.classified_df = df
                        st.rerun()

    # All transactions with inline Shared/Direct buttons — green when selected
    st.markdown("---")
    st.subheader("🔍 All Transactions")

    # Filters
    fc1, fc2 = st.columns(2)
    with fc1:
        filter_tag = st.selectbox("Filter by Status", ["All", "Untagged only", "Shared only", "Direct only"], key="shared_filter")
    with fc2:
        all_heads = ["All"] + sorted(df[df["head"].notna()]["head"].unique().tolist())
        filter_head = st.selectbox("Filter by Head", all_heads, key="shared_head_filter")

    show_df = df.copy()
    if filter_tag == "Untagged only":
        show_df = show_df[~show_df["shared_tag"].isin(["shared", "direct"])]
    elif filter_tag == "Shared only":
        show_df = show_df[show_df["shared_tag"] == "shared"]
    elif filter_tag == "Direct only":
        show_df = show_df[show_df["shared_tag"] == "direct"]
    if filter_head != "All":
        show_df = show_df[show_df["head"] == filter_head]
    if search_q3:
        show_df = search_filter(show_df, search_q3)

    # Sort: untagged first, then by head
    show_df["_sort_tagged"] = show_df["shared_tag"].isin(["shared", "direct"]).astype(int)
    show_df = show_df.sort_values(["_sort_tagged", "head", "sub_head", "name"]).drop(columns=["_sort_tagged"])

    # Pagination for speed
    PAGE_SIZE = 25
    if "shared_page" not in st.session_state:
        st.session_state.shared_page = 0
    total_pages = max(1, (len(show_df) + PAGE_SIZE - 1) // PAGE_SIZE)
    st.session_state.shared_page = min(st.session_state.shared_page, total_pages - 1)

    pc1, pc2, pc3 = st.columns([1, 2, 1])
    with pc1:
        if st.button("◀ Prev", disabled=(st.session_state.shared_page == 0), key="pg_prev"):
            st.session_state.shared_page -= 1
            st.rerun()
    with pc2:
        st.markdown(f"<div style='text-align:center'>Page {st.session_state.shared_page + 1} of {total_pages} ({len(show_df)} transactions)</div>", unsafe_allow_html=True)
    with pc3:
        if st.button("Next ▶", disabled=(st.session_state.shared_page >= total_pages - 1), key="pg_next"):
            st.session_state.shared_page += 1
            st.rerun()

    start = st.session_state.shared_page * PAGE_SIZE
    page_df = show_df.iloc[start:start + PAGE_SIZE]

    for idx, row in page_df.iterrows():
        tag = row.get("shared_tag", "")
        is_shared = (tag == "shared")
        is_direct = (tag == "direct")
        memo_text = row["memo"][:40] if row["memo"] else "—"
        currency = row["currency"] if "currency" in row.index else "USD"

        c1, ca, c2, c3, c4 = st.columns([3, 0.8, 0.8, 0.8, 0.3])
        with c1:
            st.caption(f"{row.get('qbo_account', '')} | {row.get('head', '')} | {row.get('sub_head', '')} | {memo_text}")
        with ca:
            _amt = st.number_input("Amt", value=float(df.at[idx, "amount_lcy"]), step=0.01, format="%.2f", key=f"amt_s3_{idx}", label_visibility="collapsed")
            if abs(_amt - df.at[idx, "amount_lcy"]) > 0.001:
                df.at[idx, "amount_lcy"] = _amt
                if st.session_state.raw_df is not None and idx in st.session_state.raw_df.index:
                    st.session_state.raw_df.at[idx, "amount_lcy"] = _amt
                st.session_state.classified_df = df
        with c2:
            btn_type_s = "primary" if is_shared else "secondary"
            if st.button("Shared", key=f"tag_shared_{idx}", use_container_width=True, type=btn_type_s):
                desc_key = row["description"].strip().lower()
                shared_rules[desc_key] = "shared"
                df.at[idx, "shared_tag"] = "shared"
                save_shared_rules(shared_rules)
                st.session_state.classified_df = df
        with c3:
            btn_type_d = "primary" if is_direct else "secondary"
            if st.button("Direct", key=f"tag_direct_{idx}", use_container_width=True, type=btn_type_d):
                desc_key = row["description"].strip().lower()
                shared_rules[desc_key] = "direct"
                df.at[idx, "shared_tag"] = "direct"
                save_shared_rules(shared_rules)
                st.session_state.classified_df = df
        with c4:
            if st.button("🗑️", key=f"del_s3_{idx}", help="Delete transaction"):
                delete_transactions([idx])
                st.rerun()

    st.markdown("---")
    remaining = df[~df["shared_tag"].isin(["shared", "direct"])]
    if remaining.empty:
        st.success("🎉 All transactions tagged!")
    else:
        st.info(f"ℹ️ {len(remaining)} transactions still untagged. They'll be treated as Direct on submit.")

    st.session_state.classified_df = df

    if st.button("▶️ Proceed to Apportionment", type="primary", use_container_width=True):
        # Default untagged to direct before proceeding
        df.loc[~df["shared_tag"].isin(["shared", "direct"]), "shared_tag"] = "direct"
        st.session_state.classified_df = df
        st.session_state.step = 4
        st.rerun()


# ============================================================
# STEP 4: Apportionment Percentages (Transaction-Level)
# ============================================================
elif st.session_state.step == 4:
    st.title("📊 Step 4: Set Apportionment Percentages")
    st.markdown("Set **%** split for each shared transaction across **Private Label brands**. Each transaction must total 100%.")

    add_transaction_form()

    search_q4 = st.text_input("🔍 Search transactions", placeholder="Type to filter by name, memo, head, brand...", key="search_s4")

    df = st.session_state.classified_df
    txn_apportionment = load_txn_apportionment()

    shared_df = df[df["shared_tag"] == "shared"].copy()
    if shared_df.empty:
        st.info("No shared expenses found. All transactions are Direct.")
        if st.button("▶️ Proceed to FX Rates", type="primary", use_container_width=True):
            st.session_state.step = 5
            st.rerun()
    else:
        total_shared = len(shared_df)
        apportioned_count = sum(1 for idx in shared_df.index if str(idx) in txn_apportionment)
        unapportioned_count = total_shared - apportioned_count

        # Summary tracker
        col1, col2, col3 = st.columns(3)
        col1.metric("✅ Apportioned", apportioned_count)
        col2.metric("❓ Unapportioned", unapportioned_count)
        col3.metric("🔄 Total Shared", total_shared)

        # Brand short names for column headers
        BRAND_SHORT = {
            "Fomin LLC": "Fomin",
            "Kaiya": "Kaiya",
            "Luna Naturals LLC": "Luna",
            "Paper Party LLC": "Paper P",
            "Rockport Tools": "Rockport",
            "Roofus Pet LLC": "Roofus",
            "Soul Mama": "Soul M",
        }
        SHORT_TO_FULL = {v: k for k, v in BRAND_SHORT.items()}
        brand_cols = [BRAND_SHORT.get(b, b) for b in PL_BRANDS]
        default_pct = round(100.0 / len(PL_BRANDS), 2)

        # ---- Quick Apply: Same % to All ----
        st.subheader("⚡ Apply Same % to All Transactions")
        apply_all = st.checkbox("Set one split for all shared transactions", key="apply_all_txn")
        if apply_all:
            all_splits = {}
            cols = st.columns(len(PL_BRANDS))
            for i, brand in enumerate(PL_BRANDS):
                with cols[i]:
                    all_splits[brand] = st.number_input(
                        BRAND_SHORT.get(brand, brand),
                        min_value=0.0, max_value=100.0,
                        value=default_pct,
                        step=0.5, key=f"all_txn_pct_{brand}",
                    )
            total_all = sum(all_splits.values())
            if abs(total_all - 100.0) > 0.01:
                st.warning(f"Total = {total_all:.2f}% (must be 100%)")
            else:
                st.success(f"Total = {total_all:.2f}%")
            if st.button("💾 Apply to All Transactions", key="save_all_txn"):
                if abs(total_all - 100.0) <= 0.01:
                    for idx in shared_df.index:
                        txn_apportionment[str(idx)] = all_splits.copy()
                    save_txn_apportionment(txn_apportionment)
                    st.success(f"Applied same split to all {total_shared} transactions!")
                    st.rerun()

        # ---- Quick Apply by Head ----
        shared_heads = sorted(shared_df[shared_df["head"].notna()]["head"].unique())
        st.subheader("⚡ Apply Same % by Head")
        apply_head = st.selectbox("Select Head to bulk-apply", ["-- Select --"] + shared_heads, key="bulk_head_select")
        if apply_head != "-- Select --":
            head_splits = {}
            head_txns = shared_df[shared_df["head"] == apply_head]
            st.caption(f"{len(head_txns)} transactions under '{apply_head}'")
            cols = st.columns(len(PL_BRANDS))
            for i, brand in enumerate(PL_BRANDS):
                with cols[i]:
                    head_splits[brand] = st.number_input(
                        BRAND_SHORT.get(brand, brand),
                        min_value=0.0, max_value=100.0,
                        value=default_pct,
                        step=0.5, key=f"head_pct_{apply_head}_{brand}",
                    )
            total_head = sum(head_splits.values())
            if abs(total_head - 100.0) > 0.01:
                st.warning(f"Total = {total_head:.2f}% (must be 100%)")
            else:
                st.success(f"Total = {total_head:.2f}%")
            if st.button(f"💾 Apply to All '{apply_head}' Transactions", key="save_head_bulk"):
                if abs(total_head - 100.0) <= 0.01:
                    for idx in head_txns.index:
                        txn_apportionment[str(idx)] = head_splits.copy()
                    save_txn_apportionment(txn_apportionment)
                    st.success(f"Applied to {len(head_txns)} transactions!")
                    st.rerun()

        # ---- Per-Transaction Apportionment ----
        st.markdown("---")
        st.subheader("📊 Per-Transaction Apportionment")

        # Filters
        fc1, fc2 = st.columns(2)
        with fc1:
            filter_status = st.selectbox("Filter by Status", ["All", "Unapportioned", "Apportioned"], key="app_status_filter")
        with fc2:
            all_heads_list = ["All"] + sorted(shared_df[shared_df["head"].notna()]["head"].unique().tolist())
            filter_head = st.selectbox("Filter by Head", all_heads_list, key="app_head_filter")

        show_df = shared_df.copy()
        if filter_head != "All":
            show_df = show_df[show_df["head"] == filter_head]
        if filter_status == "Unapportioned":
            show_df = show_df[~show_df.index.map(lambda x: str(x) in txn_apportionment)]
        elif filter_status == "Apportioned":
            show_df = show_df[show_df.index.map(lambda x: str(x) in txn_apportionment)]
        if search_q4:
            show_df = search_filter(show_df, search_q4)

        # Sort: unapportioned first, then by head
        show_df["_app_done"] = show_df.index.map(lambda x: 1 if str(x) in txn_apportionment else 0)
        show_df = show_df.sort_values(["_app_done", "head", "sub_head", "name"]).drop(columns=["_app_done"])

        # Pagination
        PAGE_SIZE = 15
        if "apportion_page" not in st.session_state:
            st.session_state.apportion_page = 0
        total_pages = max(1, (len(show_df) + PAGE_SIZE - 1) // PAGE_SIZE)
        st.session_state.apportion_page = min(st.session_state.apportion_page, total_pages - 1)

        pc1, pc2, pc3 = st.columns([1, 2, 1])
        with pc1:
            if st.button("◀ Prev", disabled=(st.session_state.apportion_page == 0), key="app_pg_prev"):
                st.session_state.apportion_page -= 1
                st.rerun()
        with pc2:
            st.markdown(f"<div style='text-align:center'>Page {st.session_state.apportion_page + 1} of {total_pages} ({len(show_df)} transactions)</div>", unsafe_allow_html=True)
        with pc3:
            if st.button("Next ▶", disabled=(st.session_state.apportion_page >= total_pages - 1), key="app_pg_next"):
                st.session_state.apportion_page += 1
                st.rerun()

        start = st.session_state.apportion_page * PAGE_SIZE
        page_df = show_df.iloc[start:start + PAGE_SIZE]

        # Build editable DataFrame with brand % columns + status
        edit_rows = []
        for idx, row in page_df.iterrows():
            splits = txn_apportionment.get(str(idx), {})
            row_data = {
                "index": idx,
                "Status": "✅" if str(idx) in txn_apportionment else "❓",
                "Head": row.get("head", ""),
                "Sub-head": row.get("sub_head", ""),
                "Name": str(row.get("name", ""))[:30],
                "Memo": str(row.get("memo", ""))[:30],
                "Amount": row["amount_lcy"],
                "🗑️": False,
            }
            for brand in PL_BRANDS:
                short = BRAND_SHORT.get(brand, brand)
                row_data[short] = float(splits.get(brand, 0.0))
            edit_rows.append(row_data)

        if edit_rows:
            edit_df = pd.DataFrame(edit_rows)

            col_config = {
                "index": None,
                "Status": st.column_config.TextColumn("", disabled=True, width="small"),
                "Head": st.column_config.TextColumn("Head", disabled=True, width="small"),
                "Sub-head": st.column_config.TextColumn("Sub-head", disabled=True, width="small"),
                "Name": st.column_config.TextColumn("Name", disabled=True, width="small"),
                "Memo": st.column_config.TextColumn("Memo", disabled=True, width="small"),
                "Amount": st.column_config.NumberColumn("Amount", format="$%.2f", width="small"),
                "🗑️": st.column_config.CheckboxColumn("🗑️", default=False),
            }
            for brand in PL_BRANDS:
                short = BRAND_SHORT.get(brand, brand)
                col_config[short] = st.column_config.NumberColumn(
                    short, min_value=0.0, max_value=100.0, step=0.5, width="small",
                )

            edited = st.data_editor(
                edit_df,
                column_config=col_config,
                column_order=["Status", "Head", "Sub-head", "Name", "Memo", "Amount"] + brand_cols + ["🗑️"],
                use_container_width=True,
                height=min(600, 60 + len(edit_rows) * 38),
                num_rows="fixed",
                key="apportion_editor",
            )

            # Separate delete button — works without needing Save
            _del_checked = [erow["index"] for _, erow in edited.iterrows() if erow.get("🗑️", False)]
            if _del_checked:
                if st.button(f"🗑️ Delete {len(_del_checked)} Selected Transaction(s)", key="del_apportion_btn"):
                    delete_transactions(_del_checked)
                    st.rerun()

            if st.button("💾 Save Apportionments", key="save_apportion_btn", type="primary"):
                errors = []
                saved = 0
                for _, erow in edited.iterrows():
                    orig_idx = erow["index"]
                    # Write back amount changes
                    new_amount = float(erow["Amount"])
                    if abs(new_amount - df.at[orig_idx, "amount_lcy"]) > 0.001:
                        df.at[orig_idx, "amount_lcy"] = new_amount
                        if st.session_state.raw_df is not None and orig_idx in st.session_state.raw_df.index:
                            st.session_state.raw_df.at[orig_idx, "amount_lcy"] = new_amount
                        st.session_state.classified_df = df
                    splits = {}
                    total = 0
                    for brand in PL_BRANDS:
                        short = BRAND_SHORT.get(brand, brand)
                        pct = float(erow[short])
                        splits[brand] = round(pct, 2)
                        total += pct
                    if total < 0.01:
                        continue  # all zeros — skip
                    if abs(total - 100.0) > 0.01:
                        errors.append(f"{erow['Head']} / {erow['Name']}: total = {total:.1f}%")
                    else:
                        txn_apportionment[str(orig_idx)] = splits
                        saved += 1
                save_txn_apportionment(txn_apportionment)
                if errors:
                    st.error(f"⚠️ {len(errors)} rows don't total 100%:\n" + "\n".join(errors[:5]))
                if saved:
                    st.success(f"✅ Saved {saved} transaction(s)!")
                    st.rerun()
                elif not errors:
                    st.info("No changes to save (fill in % values first).")
        else:
            st.info("No transactions match the current filter.")

        # Bottom summary
        st.markdown("---")
        apportioned_now = sum(1 for idx in shared_df.index if str(idx) in txn_apportionment)
        if apportioned_now == total_shared:
            st.success(f"🎉 All {total_shared} shared transactions apportioned!")
        else:
            st.info(f"{apportioned_now}/{total_shared} transactions apportioned. Complete the rest before proceeding.")

        if st.button("▶️ Proceed to FX Rates", type="primary", use_container_width=True):
            st.session_state.step = 5
            st.rerun()


# ============================================================
# STEP 5: FX Rates
# ============================================================
elif st.session_state.step == 5:
    st.title("💱 Step 5: Confirm FX Rates (CAD → USD)")

    add_transaction_form()

    df = st.session_state.classified_df
    fx_rates = load_fx_rates()

    cad_df = df[df["currency"] == "CAD"]
    if cad_df.empty:
        st.info("No CAD transactions found. Skipping FX conversion.")
        if st.button("▶️ Proceed to Export", type="primary", use_container_width=True):
            st.session_state.step = 6
            st.rerun()
    else:
        required_months = get_required_months(df)
        st.markdown(f"**CAD transactions found:** {len(cad_df)} across {len(required_months)} month(s)")
        st.markdown("*Rate = CAD per 1 USD. Source: [x-rates.com](https://www.x-rates.com/average/?from=USD&to=CAD&amount=1&year=2026) monthly average*")
        st.caption("Jan 2026 avg: 1.3787 | Feb 2026 avg: 1.3647 (partial)")

        updated_rates = {}
        for month in required_months:
            current_rate = fx_rates.get(month, 1.3787)
            col1, col2 = st.columns([1, 2])
            with col1:
                st.text(f"Month: {month}")
            with col2:
                updated_rates[month] = st.number_input(
                    f"CAD/USD rate for {month}",
                    min_value=0.5,
                    max_value=3.0,
                    value=float(current_rate),
                    step=0.0001,
                    format="%.4f",
                    key=f"fx_{month}",
                )

        if st.button("💾 Save FX Rates"):
            for m, r in updated_rates.items():
                fx_rates[m] = r
            save_fx_rates(fx_rates)
            st.success("FX rates saved!")

        st.markdown("---")
        if st.button("▶️ Proceed to Export", type="primary", use_container_width=True):
            # Save rates before proceeding
            for m, r in updated_rates.items():
                fx_rates[m] = r
            save_fx_rates(fx_rates)
            st.session_state.step = 6
            st.rerun()


# ============================================================
# STEP 6: Export CSV
# ============================================================
elif st.session_state.step == 6:
    st.title("📤 Step 6: Generate & Export CSV")

    add_transaction_form()

    search_q6 = st.text_input("🔍 Search transactions", placeholder="Type to filter by name, memo, head, brand...", key="search_s6")

    # Allow last-minute deletion of source transactions before export
    with st.expander("🗑️ Remove Transactions Before Export", expanded=False):
        _src = st.session_state.classified_df
        _del_df = _src[["head", "sub_head", "name", "brand", "amount_lcy", "shared_tag"]].copy()
        _del_df = _del_df.reset_index()
        _del_df["🗑️"] = False
        _del_edited = st.data_editor(
            _del_df,
            column_config={
                "index": None,
                "🗑️": st.column_config.CheckboxColumn("🗑️", default=False),
                "head": st.column_config.TextColumn("Head", disabled=True),
                "sub_head": st.column_config.TextColumn("Sub-head", disabled=True),
                "name": st.column_config.TextColumn("Name", disabled=True),
                "brand": st.column_config.TextColumn("Brand", disabled=True),
                "amount_lcy": st.column_config.NumberColumn("Amount", format="$%.2f"),
                "shared_tag": st.column_config.TextColumn("Tag", disabled=True),
            },
            use_container_width=True,
            height=300,
            num_rows="fixed",
            key="export_del_editor",
        )
        _to_del = [r["index"] for _, r in _del_edited.iterrows() if r.get("🗑️", False)]
        # Save amount changes
        _amt_changes = 0
        for _, r in _del_edited.iterrows():
            _oi = r["index"]
            _na = float(r["amount_lcy"])
            if abs(_na - _src.at[_oi, "amount_lcy"]) > 0.001:
                st.session_state.classified_df.at[_oi, "amount_lcy"] = _na
                if st.session_state.raw_df is not None and _oi in st.session_state.raw_df.index:
                    st.session_state.raw_df.at[_oi, "amount_lcy"] = _na
                _amt_changes += 1
        if _amt_changes:
            save_state_to_disk()
            st.success(f"✅ Updated {_amt_changes} amount(s)")
        if _to_del:
            if st.button(f"🗑️ Delete {len(_to_del)} Selected Transaction(s)", type="primary"):
                delete_transactions(_to_del)
                st.rerun()

    df = st.session_state.classified_df.copy()
    fx_rates = load_fx_rates()
    txn_apportionment = load_txn_apportionment()

    with st.spinner("Building final output..."):
        output_rows = []
        sr_counter = 0

        for idx, row in df.iterrows():
            amount_lcy = row["amount_lcy"]
            currency = row["currency"]
            head = row.get("head", "") or ""
            sub_head = row.get("sub_head", "") or ""
            shared_tag = row.get("shared_tag", "direct")
            brand = row["brand"]
            account_type = row["account_type"]
            date = row["date"]

            # FX conversion
            if currency == "CAD":
                amount_usd, fx_rate = convert_cad_to_usd(amount_lcy, date, fx_rates)
                lcy_display = amount_lcy
                fx_display = fx_rate
            else:
                amount_usd = amount_lcy
                lcy_display = None
                fx_display = None

            # Month format
            if hasattr(date, "strftime"):
                month_str = date.strftime("%b-%y")
                date_val = date  # keep actual date for Excel formatting
            else:
                month_str = ""
                date_val = date

            # For payroll, concatenate Name - Memo; otherwise use memo/description
            if head == "Payroll":
                name_val = (row.get("name", "") or "").strip()
                memo_val = (row.get("memo", "") or row.get("description", "")).strip()
                parts = [p for p in [name_val, memo_val] if p]
                description = " - ".join(parts) if parts else ""
            else:
                description = row.get("memo", "") or row.get("description", "")

            if shared_tag == "shared" and head:
                # Use transaction-level apportionment
                splits = txn_apportionment.get(str(idx), None)
                if splits:
                    allocated = apportion_by_splits(amount_usd, splits)
                else:
                    allocated = apportion_by_splits(amount_usd, get_default_apportionment())
                for alloc_brand, alloc_amount in allocated:
                    sr_counter += 1
                    # Split LCY proportionally (same ratio as USD split)
                    if currency == "CAD" and lcy_display and amount_usd != 0:
                        alloc_lcy = round(lcy_display * (alloc_amount / amount_usd), 2)
                    else:
                        alloc_lcy = None
                    output_rows.append({
                        "Brand": alloc_brand,
                        "Month": month_str,
                        "Sr.": sr_counter,
                        "Account": "4. OPEX",
                        "Date": date_val,
                        "Head": head,
                        "Sub-head": sub_head,
                        "Description": description,
                        "LCY": alloc_lcy if currency == "CAD" else None,
                        "Exchange rate": fx_display if currency == "CAD" else None,
                        "AMOUNT USD": alloc_amount,
                    })
            else:
                # Direct expense — stays with the brand
                sr_counter += 1
                output_rows.append({
                    "Brand": brand,
                    "Month": month_str,
                    "Sr.": sr_counter,
                    "Account": "4. OPEX",
                    "Date": date_val,
                    "Head": head,
                    "Sub-head": sub_head,
                    "Description": description,
                    "LCY": lcy_display if currency == "CAD" else "",
                    "Exchange rate": fx_display if currency == "CAD" else "",
                    "AMOUNT USD": amount_usd,
                })

        output_df = pd.DataFrame(output_rows)

    # Display results
    st.success(f"✅ Generated {len(output_df)} output rows from {len(df)} source transactions")

    col1, col2, col3 = st.columns(3)
    col1.metric("Output Rows", len(output_df))
    col2.metric("Heads", output_df["Head"].nunique() if not output_df.empty else 0)
    col3.metric("Total USD", f"${output_df['AMOUNT USD'].sum():,.2f}" if not output_df.empty else "$0")

    # Summary by Head → Sub-head (collapsible)
    st.subheader("Summary by Head")
    if not output_df.empty:
        head_totals = output_df.groupby("Head")["AMOUNT USD"].sum().sort_values(key=abs, ascending=False)
        for head_name, head_total in head_totals.items():
            sub_summary = output_df[output_df["Head"] == head_name].groupby("Sub-head")["AMOUNT USD"].sum().sort_values(key=abs, ascending=False)
            with st.expander(f"{head_name} — ${head_total:,.2f}"):
                for sub_name, sub_total in sub_summary.items():
                    st.caption(f"{sub_name}: ${sub_total:,.2f}")

    # Preview
    st.subheader("Preview")
    if not output_df.empty:
        head_filter = st.selectbox("Filter by Head", ["All"] + sorted(output_df["Head"].dropna().unique().tolist()))
        display_df = output_df if head_filter == "All" else output_df[output_df["Head"] == head_filter]
        if search_q6:
            _s_cols = [c for c in ["Brand", "Head", "Sub-head", "Description"] if c in display_df.columns]
            _s_mask = pd.Series(True, index=display_df.index)
            for _t in search_q6.strip().lower().split():
                _t_mask = pd.Series(False, index=display_df.index)
                for _c in _s_cols:
                    _t_mask |= display_df[_c].astype(str).str.lower().str.contains(_t, na=False, regex=False)
                _s_mask &= _t_mask
            display_df = display_df[_s_mask]
        st.dataframe(display_df, use_container_width=True, height=400)

    # Export buttons
    st.subheader("Export")
    col1, col2 = st.columns(2)

    with col1:
        if not output_df.empty:
            csv_data = output_df.to_csv(index=False)
            st.download_button(
                "⬇️ Download CSV",
                csv_data,
                file_name=f"expense_report_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True,
            )

    with col2:
        if not output_df.empty:
            excel_path = f"exports/expense_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
            os.makedirs("exports", exist_ok=True)
            from openpyxl.utils import get_column_letter
            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                output_df.to_excel(writer, index=False, sheet_name="Expense Report")
                ws = writer.sheets["Expense Report"]
                # Find Date column and apply mmm-yy format
                date_col_idx = list(output_df.columns).index("Date") + 1
                date_letter = get_column_letter(date_col_idx)
                for row in range(2, ws.max_row + 1):
                    cell = ws[f"{date_letter}{row}"]
                    cell.number_format = "MMM-YY"
            with open(excel_path, "rb") as f:
                st.download_button(
                    "⬇️ Download Excel",
                    f.read(),
                    file_name=os.path.basename(excel_path),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

# ============================================================
# AUTO-SAVE STATE TO DISK (runs on every rerun)
# ============================================================
save_state_to_disk()
